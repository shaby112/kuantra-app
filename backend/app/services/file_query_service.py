import duckdb
import polars as pl
import os
import tarfile
from typing import List, Dict, Any
import logging
from openpyxl import load_workbook

import re

from app.core.config import settings

logger = logging.getLogger(__name__)

UPLOAD_DIR = settings.UPLOAD_DIR

def sanitize_table_name(name: str) -> str:
    """Sanitizes strings to be used as SQL table names."""
    # Remove extension if present
    name = os.path.splitext(name)[0]
    # Replace non-alphanumeric with underscores
    name = re.sub(r'[^a-zA-Z0-9]', '_', name)
    # Ensure it starts with a letter
    if name and name[0].isdigit():
        name = "t_" + name
    return name.lower() or "dataset"

class FileQueryService:
    @staticmethod
    def _extract_tar(file_path: str) -> str:
        """Extracts tar.gz and returns the path to the first queryable file found."""
        target_dir = file_path + "_extracted"
        if not os.path.exists(target_dir):
            os.makedirs(target_dir)
            with tarfile.open(file_path, "r:gz") as tar:
                tar.extractall(path=target_dir)
        
        # Look for valid files in the extraction
        for root, _, files in os.walk(target_dir):
            for f in files:
                if f.lower().endswith(('.csv', '.parquet', '.xlsx', '.xls')):
                    return os.path.join(root, f)
        return target_dir

    @staticmethod
    def query_file(file_path: str, sql_query: str) -> List[Dict[str, Any]]:
        """
        Queries a CSV, Excel, Parquet, or tar.gz file using DuckDB.
        """
        try:
            # Determine file type
            base_name = os.path.basename(file_path)
            table_name = sanitize_table_name(base_name)
            
            ext = os.path.splitext(file_path)[1].lower()
            if file_path.endswith('.tar.gz'):
                file_path = FileQueryService._extract_tar(file_path)
                ext = os.path.splitext(file_path)[1].lower()
                table_name = sanitize_table_name(os.path.basename(file_path))

            # Create a DuckDB connection
            con = duckdb.connect(database=':memory:')
            
            # Register the table based on file type
            if ext == '.csv':
                con.execute(f"CREATE TABLE {table_name} AS SELECT * FROM read_csv_auto('{file_path}')")
            elif ext in ['.xlsx', '.xls']:
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                try:
                    for sheet in workbook.sheetnames:
                        df = pl.read_excel(file_path, sheet_name=sheet)
                        con.register(sanitize_table_name(sheet), df)
                finally:
                    workbook.close()
            elif ext == '.parquet':
                con.execute(f"CREATE TABLE {table_name} AS SELECT * FROM read_parquet('{file_path}')")
            else:
                raise ValueError(f"Unsupported file extension: {ext}")
            
            # Execute the query
            result = con.execute(sql_query)
            rows = result.fetchall()
            columns = [desc[0] for desc in (result.description or [])]
            
            # Convert to list of dicts
            return [dict(zip(columns, row)) for row in rows]
        except Exception as e:
            logger.error(f"Error querying file {file_path}: {e}")
            raise e
        finally:
            if 'con' in locals():
                con.close()

    @staticmethod
    def get_file_schema(file_path: str) -> List[Dict[str, Any]]:
        """
        Returns the schema of the file (potentially multiple tables for Excel).
        """
        try:
            if file_path.endswith('.tar.gz'):
                file_path = FileQueryService._extract_tar(file_path)
            
            base_name = os.path.basename(file_path)
            primary_table = sanitize_table_name(base_name)
            ext = os.path.splitext(file_path)[1].lower()
            con = duckdb.connect(database=':memory:')
            
            tables_to_return = []

            if ext == '.csv':
                con.execute(f"CREATE TABLE {primary_table} AS SELECT * FROM read_csv_auto('{file_path}') LIMIT 0")
                tables_to_return.append(primary_table)
            elif ext in ['.xlsx', '.xls']:
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                try:
                    for sheet in workbook.sheetnames:
                        s_name = sanitize_table_name(sheet)
                        df = pl.read_excel(file_path, sheet_name=sheet, n_rows=0)
                        con.register(s_name, df)
                        tables_to_return.append(s_name)
                finally:
                    workbook.close()
            elif ext == '.parquet':
                con.execute(f"CREATE TABLE {primary_table} AS SELECT * FROM read_parquet('{file_path}') LIMIT 0")
                tables_to_return.append(primary_table)
            
            final_schema = []
            for t in tables_to_return:
                schema_info = con.execute(f"DESCRIBE {t}").fetchall()
                cols = [{"column_name": row[0], "column_type": row[1]} for row in schema_info]
                # Map column_type to standard naming for consistency if needed, but for now just pass as is
                final_schema.append({
                    "table": t,
                    "columns": cols
                })
            
            return final_schema
        except Exception as e:
            logger.error(f"Error getting schema for {file_path}: {e}")
            raise e
        finally:
            if 'con' in locals():
                con.close()
