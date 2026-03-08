"""
File Source Connector.

Features:
- Polars for fast file reading (10-100x faster than pandas)
- Supports CSV, Excel, Parquet
- Lazy evaluation for large files
"""

import os
import polars as pl
from typing import Dict, Any, List
from pathlib import Path

from app.core.logging import logger


def sanitize_table_name(name: str) -> str:
    """Sanitizes strings to be used as SQL table names."""
    import re
    
    # Remove extension if present
    name = os.path.splitext(name)[0]
    # Replace non-alphanumeric with underscores
    name = re.sub(r'[^a-zA-Z0-9]', '_', name)
    # Ensure it starts with a letter
    if name and name[0].isdigit():
        name = "t_" + name
    return name.lower() or "dataset"


class FileSource:
    """File data source connector using Polars."""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.extension = self._get_extension()
    
    def _get_extension(self) -> str:
        """Get file extension, handling tar.gz."""
        if self.file_path.endswith('.tar.gz'):
            return '.tar.gz'
        return os.path.splitext(self.file_path)[1].lower()
    
    def _extract_tar(self) -> str:
        """Extract tar.gz and return path to first queryable file."""
        import tarfile
        
        target_dir = self.file_path + "_extracted"
        if not os.path.exists(target_dir):
            os.makedirs(target_dir)
            with tarfile.open(self.file_path, "r:gz") as tar:
                tar.extractall(path=target_dir)
        
        # Find first valid file
        for root, _, files in os.walk(target_dir):
            for f in files:
                if f.lower().endswith(('.csv', '.parquet', '.xlsx', '.xls')):
                    return os.path.join(root, f)
        
        return target_dir
    
    def extract(self) -> Dict[str, List[Dict[str, Any]]]:
        """
        Extract data from file.
        
        Returns:
            Dict mapping table names to list of records
        """
        file_path = self.file_path
        
        # Handle tar.gz
        if self.extension == '.tar.gz':
            file_path = self._extract_tar()
            self.extension = os.path.splitext(file_path)[1].lower()
        
        base_name = os.path.basename(file_path)
        table_name = sanitize_table_name(base_name)
        
        tables_data = {}
        
        try:
            if self.extension == '.csv':
                # Use lazy scanning for large files
                lf = pl.scan_csv(file_path)
                df = lf.collect()
                if "id" not in df.columns:
                    import uuid
                    df = df.with_columns(pl.Series("id", [str(uuid.uuid4()) for _ in range(len(df))]))
                
                tables_data[table_name] = df.to_dicts()
                logger.info(f"Extracted {len(tables_data[table_name])} rows from CSV")
                
            elif self.extension in ['.xlsx', '.xls']:
                # Excel files - read all sheets
                import openpyxl
                
                wb = openpyxl.load_workbook(file_path, read_only=True)
                for sheet_name in wb.sheetnames:
                    # Polars doesn't support Excel directly, use openpyxl + conversion
                    df = pl.read_excel(file_path, sheet_name=sheet_name)
                    if "id" not in df.columns:
                        import uuid
                        df = df.with_columns(pl.Series("id", [str(uuid.uuid4()) for _ in range(len(df))]))
                    
                    sheet_table = sanitize_table_name(sheet_name)
                    tables_data[sheet_table] = df.to_dicts()
                    logger.info(f"Extracted {len(tables_data[sheet_table])} rows from sheet {sheet_name}")
                
            elif self.extension == '.parquet':
                # Use lazy scanning for Parquet
                lf = pl.scan_parquet(file_path)
                df = lf.collect()
                tables_data[table_name] = df.to_dicts()
                logger.info(f"Extracted {len(tables_data[table_name])} rows from Parquet")
                
            else:
                raise ValueError(f"Unsupported file extension: {self.extension}")
                
        except Exception as e:
            logger.error(f"Error extracting file {file_path}: {e}")
            raise
        
        return tables_data
    
    def get_schema(self) -> List[Dict[str, Any]]:
        """
        Get schema information for the file.
        
        Returns:
            List of table schemas with columns
        """
        file_path = self.file_path
        
        if self.extension == '.tar.gz':
            file_path = self._extract_tar()
            self.extension = os.path.splitext(file_path)[1].lower()
        
        base_name = os.path.basename(file_path)
        table_name = sanitize_table_name(base_name)
        
        schemas = []
        
        try:
            if self.extension == '.csv':
                # Read just first 0 rows for schema
                lf = pl.scan_csv(file_path)
                schema = lf.collect_schema()
                columns = [
                    {"column_name": name, "column_type": str(dtype)}
                    for name, dtype in schema.items()
                ]
                schemas.append({"table": table_name, "columns": columns})
                
            elif self.extension in ['.xlsx', '.xls']:
                import openpyxl
                
                wb = openpyxl.load_workbook(file_path, read_only=True)
                for sheet_name in wb.sheetnames:
                    df = pl.read_excel(file_path, sheet_name=sheet_name, n_rows=0)
                    columns = [
                        {"column_name": name, "column_type": str(dtype)}
                        for name, dtype in df.schema.items()
                    ]
                    schemas.append({
                        "table": sanitize_table_name(sheet_name),
                        "columns": columns
                    })
                
            elif self.extension == '.parquet':
                lf = pl.scan_parquet(file_path)
                schema = lf.collect_schema()
                columns = [
                    {"column_name": name, "column_type": str(dtype)}
                    for name, dtype in schema.items()
                ]
                schemas.append({"table": table_name, "columns": columns})
                
        except Exception as e:
            logger.error(f"Error getting schema for {file_path}: {e}")
            raise
        
        return schemas
